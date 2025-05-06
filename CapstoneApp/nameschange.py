import os
import random
import shutil
from openpyxl import Workbook, load_workbook

# Define paths
recognized_faces_dir = 'unrecognized_faces'
database_dir = 'database'
spreadsheet_path = 'students.xlsx'

# Create the database folder if it doesn't exist
os.makedirs(database_dir, exist_ok=True)

# Load or create the spreadsheet
if os.path.exists(spreadsheet_path):
    wb = load_workbook(spreadsheet_path)
    ws = wb.active
    
    # Find existing names (without extension) to avoid duplicates
    existing_names = set()
    for cell in ws['A'][1:]:  # Skip header
        if cell.value:
            # Handle both R-numbers and names
            existing_names.add(str(cell.value).strip())
else:
    wb = Workbook()
    ws = wb.active
    ws['A1'] = 'Student Name'
    existing_names = set()

# Process images
for filename in os.listdir(recognized_faces_dir):
    file_path = os.path.join(recognized_faces_dir, filename)
    if not os.path.isfile(file_path):
        continue

    # Get file extension
    base_name, ext = os.path.splitext(filename)
    ext = ext.lower()
    if ext not in ['.jpg', '.jpeg', '.png']:
        continue  # Skip non-image files

    # Generate a unique R-number starting with 1166
    max_attempts = 100  # Prevent infinite loops
    r_number = None
    
    for _ in range(max_attempts):
        random_part = random.randint(1000, 9999)  # 4 random digits
        candidate = f"1166{random_part}"
        if candidate not in existing_names:
            r_number = candidate
            break
    
    if r_number is None:
        print(f"Failed to generate unique R-number for {filename} after {max_attempts} attempts")
        continue

    # New file name
    new_filename = f"{r_number}{ext}"
    new_file_path = os.path.join(database_dir, new_filename)
    
    try:
        # Move and rename the file
        shutil.move(file_path, new_file_path)
        
        # Add to spreadsheet (without extension)
        ws.append([r_number])
        existing_names.add(r_number)
        print(f"Processed {filename} -> {new_filename}")
    except Exception as e:
        print(f"Error processing {filename}: {str(e)}")

# Save spreadsheet
try:
    wb.save(spreadsheet_path)
    print(f"Updated {spreadsheet_path} successfully!")
except Exception as e:
    print(f"Error saving spreadsheet: {str(e)}")

print("Processing complete.")